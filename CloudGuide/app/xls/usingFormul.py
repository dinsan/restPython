from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws["A1"] = "=SUM(12, 1)"
wb.save("formula.xlsx")


wb = Workbook("formula.xlsx", data_only=True)
ws = wb['Sheet']
cell = ws['A1']
print(cell.value)