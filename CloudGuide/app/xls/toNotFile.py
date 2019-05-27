from openpyxl import load_workbook

wb = load_workbook("xxx.xlsx", data_only=True)
ws = wb.active
ws = wb['visit spk lang of in']
max_row = ws.max_row
print(max_row)
for index in range(1, 27):

    if index > 5:
        print(index)
        ws['J' + str(index) + ''] = '=COUNTIFS($B$5:$B$30404;$J$5;$A$5:$A$30404;H' + str(index) + ')'
        wb.save("xxx.xlsx")

print("is done")
